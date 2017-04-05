# Create your views here.
# -*- coding: utf-8 -*-
#import necesarios para trabajar.

import datetime
from helpers import vista

from django.contrib.contenttypes.models import ContentType

import ho.pisa as pisa
import cStringIO as StringIO
import cgi
from django.template.loader import render_to_string 


from django.shortcuts import render_to_response
from django.http import HttpResponseRedirect, HttpResponse
from django.template import RequestContext
from grados.models import Alumno,Programa
from grados.forms import ContactForm,AddAlumnoForm,LoginForm,AlumnoForm,ProgramaForm,Miformulario,ConsultaForm,Programas

from django.utils import simplejson
from django.utils.safestring import mark_safe
#imports para la paginacion
from django.core.paginator import Paginator, InvalidPage, EmptyPage
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth import login,logout,authenticate
# entradas para paginacion
from django.core.paginator import Paginator,EmptyPage,InvalidPage
# para buscar datos
from django.db.models import Q
from reportlab.pdfgen import canvas
from io import BytesIO
from django.views.generic import ListView
from reportlab.lib.pagesizes import letter
from django.contrib.auth.decorators import login_required
import django

from reportlab.platypus.doctemplate import SimpleDocTemplate
from reportlab.platypus import Paragraph, Spacer
#from reportlab.lib import sytles
from django.db.models import F
from django.views.generic.base import TemplateView
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
#Nos devuelve un objeto resultado, en este caso un archivo de excel
from django.http.response import HttpResponse





def index_view(Request):
	return render_to_response('index.html',context_instance=RequestContext(Request))
 

def about_view(Request):
	version = django.get_version()
	mensaje =" Aplicacion Realizada en Django para el registro de Graduados del Itfip"
	ctx ={'msg':mensaje,'version':version}
	return render_to_response('acercade.html',ctx,context_instance=RequestContext(Request))


def alumnos_view(Request,pagina):
	estudiante		= 	Alumno.objects.all()
	paginator 		=	Paginator(estudiante,7) #cuantos alumnos por pagina
	try:
		page = int(pagina)		
	except:
		page = 1	
	try:
		alumnos1 = paginator.page(page)
	except:
		(EmptyPage,InvalidPage)
		alumnos1 = paginator.page(paginator.num_pages)
	ctx = {'alumnos':alumnos1}
	return render_to_response('alumnos.html',ctx,context_instance=RequestContext(Request))

def contacto_view(Request):
	info_enviado = False
	email= ""
	titulo=""
	texto=""
	if Request.method =="POST":
			formulario =	ContactForm(Request.POST)
			if formulario.is_valid():
				info_enviado =True
				email	= formulario.cleaned_data['Email']
				titulo	= formulario.cleaned_data['Titulo']
				texto	= formulario.cleaned_data['Mensaje']
				
								# configuramos el correo
				to_admin	= 'wpiraquive@itfip.edu.co'
				html_content= "informacion recibida de[%s]<br><br><br>***Mensaje<br><br>%s"%(email,texto)
				msg	= EmailMultiAlternatives('Correo de Contacto',html_content,'from@server.com',[to_admin])
				msg.attach_alternative(html_content,'text/html')# se define el contenido en html
				msg.send()# enviamos correo
				return HttpResponseRedirect('contacto.html')
	else:
			formulario	=	ContactForm()
			ctx 		=	{'form': formulario,'email': email,'titulo': titulo,'texto': texto,'info_enviado':info_enviado}
			return render_to_response('contacto.html',ctx,context_instance=RequestContext(Request))




# vista agregar alumnos

def add_alumno(Request):
	# Para depurar
	'''
	import pdb
	pdb.set_trace()
	'''
	info =	"Inicializando"
	if Request.user.is_authenticated():
	
	

		if Request.method 	== "POST":
			form =	AddAlumnoForm(Request.POST)
		
		

			if form.is_valid():
					nit 		= form.cleaned_data['nit']
					nombres		= form.cleaned_data['nombres']
					programa	= form.cleaned_data['programa']
					libro		= form.cleaned_data['libro']	
					acta 		= form.cleaned_data['acta']
					folio 		= form.cleaned_data['folio']
					fecha		= form.cleaned_data['fecha']
					diploma		= form.cleaned_data['diploma']
					asistencia	= form.cleaned_data['asistencia']

					p 			= Alumno()
					p.nit		= nit
					p.nombres	= nombres
					p.programa 	= programa
					p.libro		= libro
					p.acta 		= acta
					p.folio 	= folio
					p.fecha 	= fecha
					p.diploma	= diploma
					p.asistencia= asistencia
					p.save() # guardar la informacion

					info= "Se guardo satisfactoriamente!!!"
					ctx = {'informacion': info, 'form': form}
					#ctx = ctx  =	{'form': form}
					return render_to_response('agregaralumno.html',ctx,context_instance=RequestContext(Request))

				#ctx = {'informacion'}
			else:
				info= "informacion con datos incorrectos"	
		form = AddAlumnoForm()
		ctx  =	{'form': form}
		return render_to_response('agregaralumno.html',ctx,context_instance=RequestContext(Request))
	else:
		return HttpResponseRedirect('/')


	'''else:
		form =	AddAlumnoForm()
		ctx  =	{'form': form}




		
	return render_to_response('agregaralumno.html',ctx,context_instance=RequestContext(Request))
	#return render_to_response('agregaralumno.html',context_instance=RequestContext(Request))	'''


def login_view(request):
	
	#import pdb
	#pdb.set_trace()
	mensaje= " "
	if request.user.is_authenticated():
			return HttpResponseRedirect('/')
	else:
			if request.method == "POST":
				form 	= LoginForm(request.POST)
				if form.is_valid():
						username	=	form.cleaned_data['username']
						password	=	form.cleaned_data['password']
						usuario		=	authenticate(username=username,password=password)
						if usuario is not None and usuario.is_active:
								login(request,usuario)
								return HttpResponseRedirect('/')
						else:
								mensaje ="Usuario y/o Password Incorrectos"

			form = LoginForm()
			ctx= {'form':form,'mensaje':mensaje}				
			return render_to_response('login.html',ctx,context_instance=RequestContext(request))

def logout_view(request):
	logout(request)
	return HttpResponseRedirect('/')

def singleAlumno_view(request,id_alum):
	alum 	=	Alumno.objects.get(id=id_alum)
	ctx 	=	{'alumno':alum}
	return render_to_response('singlealumno.html',ctx,context_instance=RequestContext(request))


def edit_view(request,id_alum):
	P = Alumno.objects.get(id=id_alum)
	if request.method=="GET":
		form = AddAlumnoForm(initial={
						 'nit':P.nit,
						 'nombres':P.nombres,
						 'programa':P.programa,
						 'libro':P.libro,
						 'acta' :P.acta,
						 'folio' :P.folio,
						 'fecha' :P.fecha,
				})
		
		ctx ={'form':form,'Alumno':P}
			
		return render_to_response('editaralumno.html',ctx,context_instance=RequestContext(request))
	else:
		return HttpResponseRedirect('/')


def search(request):
    query = request.GET.get('q','').upper()
    
    if query:
    	qset = (
            Q(nombres__icontains=query) |
            Q(nit__icontains=query)|
            Q(acta__icontains=query) 
           
           
        )
        results = Alumno.objects.filter(qset).distinct()
    else:
        results = []
    return render_to_response('search.html', {"results": results,"query": query.upper()})






def search1(request):
	

    query = str(request.GET.get('q','')).upper()
    if request.method == 'POST':
    	
    	for asistencia in request.POST.getlist('asistencia'):
    		a = Alumno.objects.get(id=asistencia)
    		a.asistencia = True
    		a.save()
    		query = str(request.POST.get('q')).upper()

    
    if query:
    	qset = (
            Q(nombres__icontains=query) |
            Q(nit__icontains=query)|
            Q(acta__icontains=query) 
           
           
        )
        results = Alumno.objects.filter(qset).distinct()
    else:
        results = []
    contexto = RequestContext(request)    
    return render_to_response('search1.html', {"results": results,"query": query}, contexto)

	#return response

def Alumnos_view(request):
	if request.method =="POST":
		formulario =AlumnoForm(request.POST,request.FILES)
		if formulario.is_valid():
			formulario.save()
			return HttpResponseRedirect('alumnos1.html')
	else:
		formulario= AlumnoForm()
	return render_to_response('alumnos1.html',{'formulario':formulario},context_instance=RequestContext(request))	



def Progama_view(request):
	if request.method=='POST':
		formulario= ProgramaForm(request.POST,request.FILES)
		if formulario.is_valid():
			formulario.save()
			return HttpResponseRedirect('/')
	else:
		formulario = ProgramaForm()
	return render_to_response('programas.html',{'formulario':formulario},context_instance=RequestContext(request))
		



def pdf(f):
    """
     Decorador que convierte una vista en pdf
    """
    def vista(request, *args, **kwargs):
        response = f(request, *args, **kwargs)
        result = StringIO.StringIO()
        pdf = pisa.pisaDocument(StringIO.StringIO(response.encode("UTF-8")), result)
        if not pdf.err:
            response = HttpResponse(result.getvalue(), content_type='application/pdf')
            return response
        return HttpResponse('Error al generar el PDF: %s' % cgi.escape(response))
        
    return vista

   # Para su uso en las vistas

@pdf
def mi_vista(request):    
	'''if request.GET.has_key('nit'):'''
	'''prueba= Alumno.objects.filter( nit = request.GET['nit'] )'''
	prueba= Alumno.objects.all( )
	html=render_to_string('alumnopdf.html', {'alumnos': prueba},context_instance=RequestContext(request))
	return html



@pdf
def mi_programa(request):    
	'''if request.GET.has_key('nit'):'''
	'''prueba= Alumno.objects.filter( nit = request.GET['nit'] )'''
	prueba= Programa.objects.all( )
	html=render_to_string('programas1.html', {'programas': prueba},context_instance=RequestContext(request))
	return html

def prueba_vista(request):
	mi_form = Miformulario(user=request.user)
	return render_to_response('pruebas.html',{'mi_form':mi_form},context_instance=RequestContext(request))


def fechas_view(request):
	form = ConsultaForm()
	return render_to_response('consulta2.html', {'form': form}, context_instance=RequestContext(request) )

@pdf
def consultafecha(request):
	'''import pdb
	pdb.set_trace()'''

	fecha = request.POST.get('Fecha_Inicial')
	fecha1 = request.POST.get('Fecha_Final')
	f_inicial=fecha
	#f_inicial = datetime.datetime.strptime(fecha,"%d/%m/%Y").strftime("%Y-%m-%d")
	#f_inicial= request.POST['Fecha_Inicial'] se cambio por el dato anterior para convertir el formato de fecha del calendario
	f_cierre= fecha1
	
	prueba = Alumno.objects.filter( fecha__range=(f_inicial,f_cierre)).order_by('fecha')
	prueba.group_by=['programa']


	
	#prueba = Alumno.objects.filter(fecha__range=( Fecha_Incial, Fecha_Final ))
	html=render_to_string('alumnopdf.html', {'prueba': prueba},context_instance=RequestContext(request))
	return html

def programas2(request):
	#import pdb
	#pdb.set_trace()
	form = Programas()
	return render_to_response('programas2.html', {'form': form},context_instance=RequestContext(request))

@pdf
def reporte_programas(request):
	
	
	'''import pdb
	pdb.set_trace()'''
		
	programa= request.POST['programa']
	prueba= Alumno.objects.filter(programa=programa)
	html=render_to_string('reporte_programa.html', {'prueba': prueba},context_instance=RequestContext(request))
	return html

def delete_view(request,id_alum):
	 p = Alumno.objects.get(id=id_alum)
	 p.delete()
	 return HttpResponseRedirect('/search.html/')


def asistencia_view(request,pagina):
	estudiante		= 	Alumno.objects.all()
	paginator 		=	Paginator(estudiante,7) #cuantos alumnos por pagina
	try:
		page = int(pagina)		
	except:
		page = 1	
	try:
		alumnos1 = paginator.page(page)
	except:
		(EmptyPage,InvalidPage)
		alumnos1 = paginator.page(paginator.num_pages)
	ctx = {'alumnos':alumnos1}
	return render_to_response('asistencia.html',ctx,context_instance=RequestContext(request))


@pdf
def mi_asistencia(request):    

	prueba= Alumno.objects.filter(asistencia=1)
	html=render_to_string('alumnopdf.html', {'prueba': prueba},context_instance=RequestContext(request))
	return html

def borra_asistencias(request):
	   
    
    a = Alumno.objects.filter(asistencia=True).update( asistencia = False)
    	
    return HttpResponseRedirect('/')
  
    
# Genera exporta a excel la consulta que se haga en la pagina
    
class ReportePersonasExcel(TemplateView):
     
    #Usamos el método get para generar el archivo excel 
    def get(self, request, *args, **kwargs):
        #Obtenemos todas las personas de nuestra base de datos
        personas = Alumno.objects.all()
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws['B1'] = 'REPORTE DE GRADUADOS'
        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:E1')
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = 'DOCUMENTO'
        ws['C3'] = 'NOMBRES'
        ws['D3'] = 'PROGRAMA'
        ws['E3'] = 'ACTA'       
        ws['F3'] = 'FOLIO' 
        ws['G3'] = 'FECHA' 
        cont=6
        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for persona in personas:
            ws.cell(row=cont,column=2).value = persona.nit
            ws.cell(row=cont,column=3).value = persona.nombres
            ws.cell(row=cont,column=4).value = persona.programa
            ws.cell(row=cont,column=5).value = persona.acta
            ws.cell(row=cont,column=6).value = persona.folio
            ws.cell(row=cont,column=7).value = persona.fecha
            cont = cont + 1
        #Establecemos el nombre del archivo
        nombre_archivo ="ReportePersonasExcel.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response